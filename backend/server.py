import os
from pathlib import Path
from dotenv import load_dotenv

# Load environment variables immediately, before any other imports
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, status
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorDatabase
import logging
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import base64
import io
import openpyxl
import random

from models import (
    User, UserCreate, UserLogin, TokenResponse, RefreshTokenRequest,
    ForgotPasswordRequest, ResetPasswordRequest, College, CollegeCreate,
    Subscription, SubscriptionCreate, Student, StudentBulkUpload,
    Question, QuestionBankUpload, ExamSet, ExamSetCreate, ExamSetUpdate, VivaExam,
    QARecord, AnswerSubmission, VoiceInteraction, TranscribeRequest,
    UserRole, SubscriptionStatus, PaymentMethod, DifficultyLevel, ExamStatus,
    ExamConfig
)
from auth import (
    hash_password, verify_password, create_access_token, create_refresh_token,
    create_reset_token, verify_token, get_current_user, require_role
)
from email_service import send_password_reset_email, send_welcome_email
from voice_service import voice_service

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "divyanshgupta5748@gmail.com")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "#PSTDG5748")
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", "/app/backend/uploads"))
UPLOAD_DIR.mkdir(exist_ok=True, parents=True)

app = FastAPI(title="College Viva Voice-Agent Platform API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()


async def get_db() -> AsyncIOMotorDatabase:
    return db


@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_create: UserCreate):
    existing_user = await db.users.find_one({"email": user_create.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )
    
    user = User(
        email=user_create.email,
        password_hash=hash_password(user_create.password),
        role=user_create.role,
        full_name=user_create.full_name,
        college_id=user_create.college_id
    )
    
    user_dict = user.model_dump()
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    user_dict['updated_at'] = user_dict['updated_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    await send_welcome_email(user.email, user.full_name, user.role.value)
    
    access_token = create_access_token({"sub": user.id, "role": user.role.value})
    refresh_token = create_refresh_token({"sub": user.id})
    
    user_response = user.model_dump()
    user_response.pop("password_hash", None)
    
    return TokenResponse(
        access_token=access_token,
        refresh_token=refresh_token,
        user=user_response
    )


@api_router.post("/auth/login", response_model=TokenResponse)
async def login(user_login: UserLogin):
    try:
        user_data = await db.users.find_one({"email": user_login.email}, {"_id": 0})
        if not user_data:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid credentials"
            )
        
        user = User(**user_data)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error for {user_login.email}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Internal login error. Please contact support."
        )
    
    if not verify_password(user_login.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials"
        )
    
    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account is inactive"
        )
    
    access_token = create_access_token({"sub": user.id, "role": user.role.value})
    refresh_token = create_refresh_token({"sub": user.id})
    
    user_response = user.model_dump()
    user_response.pop("password_hash", None)
    
    if user.college_id:
        college_data = await db.colleges.find_one({"id": user.college_id}, {"_id": 0, "name": 1})
        if college_data:
            user_response["college_name"] = college_data.get("name")

    return TokenResponse(
        access_token=access_token,
        refresh_token=refresh_token,
        user=user_response
    )


@api_router.post("/auth/refresh", response_model=Dict[str, str])
async def refresh_token(refresh_req: RefreshTokenRequest):
    payload = verify_token(refresh_req.refresh_token, "refresh")
    user_id = payload.get("sub")
    
    user_data = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="User not found"
        )
    
    user = User(**user_data)
    access_token = create_access_token({"sub": user.id, "role": user.role.value})
    
    return {"access_token": access_token, "token_type": "bearer"}


@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    user_data = await db.users.find_one({"email": request.email}, {"_id": 0})
    if not user_data:
        return {"message": "If the email exists, a reset link has been sent"}
    
    reset_token = create_reset_token(request.email)
    await send_password_reset_email(request.email, reset_token)
    
    return {"message": "If the email exists, a reset link has been sent"}


@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    payload = verify_token(request.token, "reset")
    email = payload.get("email")
    
    if not email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid token"
        )
    
    new_password_hash = hash_password(request.new_password)
    
    result = await db.users.update_one(
        {"email": email},
        {"$set": {"password_hash": new_password_hash, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    return {"message": "Password reset successful"}


async def get_current_user_dep(
    auth: HTTPAuthorizationCredentials = Depends(security),
    database: AsyncIOMotorDatabase = Depends(get_db)
) -> User:
    token = auth.credentials
    payload = verify_token(token, "access")
    user_id = payload.get("sub")
    
    user_data = await database.users.find_one({"id": user_id}, {"_id": 0})
    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="User not found"
        )
    
    return User(**user_data)


@api_router.post("/colleges/create", response_model=College)
async def create_college(college_create: CollegeCreate):
    existing_college = await db.colleges.find_one({"admin_email": college_create.admin_email}, {"_id": 0})
    if existing_college:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="College with this admin email already exists"
        )
    
    user_create = UserCreate(
        email=college_create.admin_email,
        password=college_create.admin_password,
        role=UserRole.COLLEGE_ADMIN,
        full_name=college_create.admin_full_name
    )
    
    existing_user = await db.users.find_one({"email": user_create.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Admin email already registered"
        )
    
    user = User(
        email=user_create.email,
        password_hash=hash_password(user_create.password),
        role=user_create.role,
        full_name=user_create.full_name
    )
    
    user_dict = user.model_dump()
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    user_dict['updated_at'] = user_dict['updated_at'].isoformat()
    await db.users.insert_one(user_dict)
    
    is_admin = college_create.admin_email.lower() == ADMIN_EMAIL.lower()
    
    college = College(
        name=college_create.name,
        admin_email=college_create.admin_email,
        admin_id=user.id,
        subscription_status=SubscriptionStatus.ACTIVE if is_admin else SubscriptionStatus.PENDING
    )
    
    college_dict = college.model_dump()
    college_dict['created_at'] = college_dict['created_at'].isoformat()
    college_dict['updated_at'] = college_dict['updated_at'].isoformat()
    if college_dict.get('subscription_expires_at'):
        college_dict['subscription_expires_at'] = college_dict['subscription_expires_at'].isoformat()
    
    await db.colleges.insert_one(college_dict)
    
    await db.users.update_one(
        {"id": user.id},
        {"$set": {"college_id": college.id}}
    )
    
    await send_welcome_email(user.email, user.full_name, user.role.value)
    
    return college


@api_router.get("/colleges/my")
async def get_my_college(current_user: User = Depends(get_current_user_dep)):
    if current_user.role not in [UserRole.COLLEGE_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can access this"
        )
    
    college_data = await db.colleges.find_one({"admin_id": current_user.id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    return college_data



@api_router.get("/colleges/all")
async def get_all_colleges(current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.GLOBAL_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only global admin can access this"
        )
    
    colleges = await db.colleges.find({}, {"_id": 0}).to_list(1000)
    return colleges

@api_router.post("/subscriptions/create", response_model=Subscription)
async def create_subscription(
    sub_create: SubscriptionCreate,
    current_user: User = Depends(get_current_user_dep)
):
    college_data = await db.colleges.find_one({"id": sub_create.college_id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    
    college = College(**college_data)
    
    if current_user.role == UserRole.COLLEGE_ADMIN and college.admin_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized for this college"
        )
    
    plan_prices = {
        "basic": 499.0,
        "standard": 999.0,
        "premium": 1999.0
    }
    
    price = plan_prices.get(sub_create.plan_name.lower(), 999.0)
    
    subscription = Subscription(
        college_id=sub_create.college_id,
        plan_name=sub_create.plan_name,
        price=price,
        payment_method=sub_create.payment_method,
        custom_payment_details=sub_create.custom_payment_details,
        status=SubscriptionStatus.PENDING,
        admin_approved=False if sub_create.payment_method == PaymentMethod.CUSTOM else False
    )
    
    sub_dict = subscription.model_dump()
    sub_dict['created_at'] = sub_dict['created_at'].isoformat()
    if sub_dict.get('expires_at'):
        sub_dict['expires_at'] = sub_dict['expires_at'].isoformat()
    
    await db.subscriptions.insert_one(sub_dict)
    
    return subscription


@api_router.post("/subscriptions/{subscription_id}/approve")
async def approve_subscription(
    subscription_id: str,
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.GLOBAL_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only global admin can approve subscriptions"
        )
    
    sub_data = await db.subscriptions.find_one({"id": subscription_id}, {"_id": 0})
    if not sub_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Subscription not found"
        )
    
    expires_at = datetime.now(timezone.utc) + timedelta(days=365)
    
    await db.subscriptions.update_one(
        {"id": subscription_id},
        {"$set": {
            "admin_approved": True,
            "status": SubscriptionStatus.ACTIVE.value,
            "expires_at": expires_at.isoformat()
        }}
    )
    
    await db.colleges.update_one(
        {"id": sub_data["college_id"]},
        {"$set": {
            "subscription_status": SubscriptionStatus.ACTIVE.value,
            "subscription_plan": sub_data["plan_name"],
            "subscription_expires_at": expires_at.isoformat()
        }}
    )
    
    return {"message": "Subscription approved"}


@api_router.get("/subscriptions/pending")
async def get_pending_subscriptions(current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.GLOBAL_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only global admin can access this"
        )
    
    subs = await db.subscriptions.find(
        {"payment_method": PaymentMethod.CUSTOM.value, "admin_approved": False},
        {"_id": 0}
    ).to_list(1000)
    
    return subs


def excel_to_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (int, float)) and float(val).is_integer():
        return str(int(val))
    return str(val)


@api_router.post("/students/bulk-upload")
async def bulk_upload_students(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can upload students"
        )
    
    college_data = await db.colleges.find_one({"admin_id": current_user.id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    
    college = College(**college_data)
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        students_created = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1] or not row[2]:
                continue
            
            student_id = excel_to_str(row[0])
            full_name = excel_to_str(row[1])
            email = excel_to_str(row[2])
            phone = excel_to_str(row[3]) if len(row) > 3 and row[3] else None
            password_raw = excel_to_str(row[4]) if len(row) > 4 and row[4] else "Student@123"
            batch = excel_to_str(row[5]) if len(row) > 5 and row[5] else None
            
            existing_student = await db.students.find_one({
                "college_id": college.id,
                "student_id": student_id
            }, {"_id": 0})
            
            if existing_student:
                continue
            
            existing_user = await db.users.find_one({"email": email}, {"_id": 0})
            if existing_user:
                user = User(**existing_user)
            else:
                user = User(
                    email=email,
                    password_hash=hash_password(password_raw),
                    role=UserRole.STUDENT,
                    full_name=full_name,
                    college_id=college.id,
                    student_id=student_id
                )
                
                user_dict = user.model_dump()
                user_dict['created_at'] = user_dict['created_at'].isoformat()
                user_dict['updated_at'] = user_dict['updated_at'].isoformat()
                await db.users.insert_one(user_dict)
            
            student = Student(
                college_id=college.id,
                student_id=student_id,
                full_name=full_name,
                email=email,
                user_id=user.id,
                phone_number=phone,
                password=password_raw,
                batch=batch
            )
            
            student_dict = student.model_dump()
            student_dict['created_at'] = student_dict['created_at'].isoformat()
            await db.students.insert_one(student_dict)
            
            students_created += 1
        
        return {"message": f"Successfully uploaded {students_created} students"}
    
    except Exception as e:
        logger.error(f"Error uploading students: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error processing Excel file: {str(e)}"
        )


@api_router.get("/students/my-college")
async def get_college_students(current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can access this"
        )
    
    college_data = await db.colleges.find_one({"admin_id": current_user.id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    
    students = await db.students.find({"college_id": college_data["id"]}, {"_id": 0}).to_list(1000)
    return students


@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str, current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    student = await db.students.find_one({"id": student_id})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
        
    # Delete student record and associated user
    await db.students.delete_one({"id": student_id})
    await db.users.delete_one({"id": student["user_id"]})
    
    return {"message": "Student deleted"}


@api_router.post("/questions/bulk-upload")
async def bulk_upload_questions(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can upload questions"
        )
    
    college_data = await db.colleges.find_one({"admin_id": current_user.id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    
    college = College(**college_data)
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        questions_created = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1] or not row[2]:
                continue
            
            subject = str(row[0]) if row[0] else "General"
            topic = str(row[1]) if row[1] else "General"
            question_text = str(row[2])
            answer_key = str(row[3])
            difficulty_str = str(row[4]).lower() if len(row) > 4 and row[4] else "medium"
            
            if difficulty_str not in ["easy", "medium", "hard"]:
                difficulty_str = "medium"
            
            # Check for duplicates
            existing_question = await db.questions.find_one({
                "college_id": college.id,
                "question_text": question_text,
                "subject": subject,
                "topic": topic
            })
            if existing_question:
                continue

            question = Question(
                college_id=college.id,
                question_text=question_text,
                difficulty_level=DifficultyLevel(difficulty_str),
                answer_key=answer_key,
                subject=subject,
                topic=topic
            )
            
            question_dict = question.model_dump()
            question_dict['difficulty_level'] = question_dict['difficulty_level'].value
            question_dict['created_at'] = question_dict['created_at'].isoformat()
            await db.questions.insert_one(question_dict)
            
            questions_created += 1
        
        return {"message": f"Successfully uploaded {questions_created} questions"}
    
    except Exception as e:
        logger.error(f"Error uploading questions: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error processing Excel file: {str(e)}"
        )


@api_router.get("/questions/my-college")
async def get_college_questions(current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can access this"
        )
    
    college_data = await db.colleges.find_one({"admin_id": current_user.id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    
    questions = await db.questions.find({"college_id": college_data["id"]}, {"_id": 0}).to_list(1000)
    return questions


@api_router.delete("/questions/{question_id}")
async def delete_question(question_id: str, current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.questions.delete_one({"id": question_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Question not found")
    
    return {"message": "Question deleted"}


@api_router.post("/exam-sets/create", response_model=ExamSet)
async def create_exam_set(
    exam_set_create: ExamSetCreate,
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can create exam sets"
        )
    
    college_data = await db.colleges.find_one({"admin_id": current_user.id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    
    college = College(**college_data)
    
    end_time = exam_set_create.end_time
    if not end_time:
        # If end_time is not provided, calculate it based on start_time and exam duration.
        # This sets the grace period to start the exam.
        end_time = exam_set_create.start_time + timedelta(minutes=exam_set_create.exam_config.duration_minutes)

    exam_set = ExamSet(
        college_id=college.id,
        name=exam_set_create.name,
        exam_config=exam_set_create.exam_config,
        selected_subject=exam_set_create.selected_subject,
        selected_topics=exam_set_create.selected_topics,
        student_ids=exam_set_create.student_ids,
        start_time=exam_set_create.start_time,
        end_time=end_time,
        created_by=current_user.id
    )
    
    exam_set_dict = exam_set.model_dump()
    exam_set_dict['start_time'] = exam_set_dict['start_time'].isoformat()
    exam_set_dict['end_time'] = end_time.isoformat()
    exam_set_dict['created_at'] = exam_set_dict['created_at'].isoformat()
    await db.exam_sets.insert_one(exam_set_dict)
    
    for student_id in exam_set_create.student_ids:
        viva_exam = VivaExam(
            exam_set_id=exam_set.id,
            student_id=student_id,
            college_id=college.id,
            status=ExamStatus.SCHEDULED
        )
        
        viva_dict = viva_exam.model_dump()
        viva_dict['status'] = viva_dict['status'].value
        viva_dict['created_at'] = viva_dict['created_at'].isoformat()
        if viva_dict.get('started_at'):
            viva_dict['started_at'] = viva_dict['started_at'].isoformat()
        if viva_dict.get('completed_at'):
            viva_dict['completed_at'] = viva_dict['completed_at'].isoformat()
        await db.viva_exams.insert_one(viva_dict)
    
    return exam_set


@api_router.get("/exam-sets/my-college")
async def get_college_exam_sets(current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can access this"
        )
    
    college_data = await db.colleges.find_one({"admin_id": current_user.id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    
    exam_sets = await db.exam_sets.find({"college_id": college_data["id"]}, {"_id": 0}).to_list(1000)
    return exam_sets

@api_router.delete("/exam-sets/{exam_set_id}")
async def delete_exam_set(
    exam_set_id: str,
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can delete exam sets"
        )

    # Delete the exam set
    result = await db.exam_sets.delete_one({"id": exam_set_id, "college_id": current_user.college_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Exam set not found")

    # Delete associated viva exams
    await db.viva_exams.delete_many({"exam_set_id": exam_set_id})
    
    return {"message": "Exam set deleted successfully"}


@api_router.get("/exam-sets/{exam_set_id}")
async def get_exam_set_details(
    exam_set_id: str,
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can access this"
        )

    exam_set_data = await db.exam_sets.find_one({"id": exam_set_id, "college_id": current_user.college_id}, {"_id": 0})
    if not exam_set_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Exam set not found or you do not have permission to view it."
        )

    # Get student details
    student_ids = exam_set_data.get("student_ids", [])
    students_cursor = db.students.find({"id": {"$in": student_ids}}, {"_id": 0})
    students = await students_cursor.to_list(length=None)

    # Get individual viva exam statuses for this set
    viva_exams_cursor = db.viva_exams.find({"exam_set_id": exam_set_id}, {"_id": 0})
    viva_exams = await viva_exams_cursor.to_list(length=None)

    # Create a map of student_id to viva_exam for easier lookup on the frontend
    viva_exams_map = {exam["student_id"]: exam for exam in viva_exams}

    # Combine student data with their viva exam status
    student_details = []
    for student in students:
        student_detail = student.copy()
        viva_exam = viva_exams_map.get(student["id"])
        student_detail["viva_exam"] = viva_exam
        student_details.append(student_detail)

    return {
        "exam_set": exam_set_data,
        "students": student_details
    }

@api_router.patch("/exam-sets/{exam_set_id}", response_model=ExamSet)
async def update_exam_set(
    exam_set_id: str,
    exam_set_update: ExamSetUpdate,
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can update exam sets"
        )

    existing_exam_set_data = await db.exam_sets.find_one({
        "id": exam_set_id,
        "college_id": current_user.college_id
    })
    if not existing_exam_set_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Exam set not found"
        )

    # Prevent updates if any student has started the exam
    in_progress_vivas = await db.viva_exams.count_documents({
        "exam_set_id": exam_set_id,
        "status": {"$ne": ExamStatus.SCHEDULED.value}
    })
    if in_progress_vivas > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot update an exam set after it has started for any student."
        )

    update_data = exam_set_update.model_dump(exclude_unset=True)

    # Handle student_ids update
    if "student_ids" in update_data:
        old_student_ids = set(existing_exam_set_data.get("student_ids", []))
        new_student_ids = set(update_data["student_ids"])

        students_to_add = new_student_ids - old_student_ids
        students_to_remove = old_student_ids - new_student_ids

        # Add new viva exams for new students
        for student_id in students_to_add:
            viva_exam = VivaExam(
                exam_set_id=exam_set_id,
                student_id=student_id,
                college_id=current_user.college_id,
                status=ExamStatus.SCHEDULED
            )
            viva_dict = viva_exam.model_dump()
            viva_dict['status'] = viva_dict['status'].value
            viva_dict['created_at'] = viva_dict['created_at'].isoformat()
            if viva_dict.get('started_at'):
                viva_dict['started_at'] = viva_dict['started_at'].isoformat()
            if viva_dict.get('completed_at'):
                viva_dict['completed_at'] = viva_dict['completed_at'].isoformat()
            await db.viva_exams.insert_one(viva_dict)

        # Remove viva exams for removed students
        if students_to_remove:
            await db.viva_exams.delete_many({
                "exam_set_id": exam_set_id,
                "student_id": {"$in": list(students_to_remove)}
            })

    # Handle time updates
    final_start_time_str = update_data.get("start_time", existing_exam_set_data["start_time"])
    final_start_time = datetime.fromisoformat(final_start_time_str) if isinstance(final_start_time_str, str) else final_start_time_str

    if "exam_config" in update_data:
        final_exam_config = ExamConfig(**update_data["exam_config"])
        update_data["exam_config"] = final_exam_config.model_dump() # ensure it's a dict for mongo
    else:
        final_exam_config = ExamConfig(**existing_exam_set_data["exam_config"])

    if "end_time" not in update_data:
        update_data["end_time"] = final_start_time + timedelta(minutes=final_exam_config.duration_minutes)

    for key in ["start_time", "end_time"]:
        if key in update_data and isinstance(update_data[key], datetime):
            update_data[key] = update_data[key].isoformat()

    if update_data:
        await db.exam_sets.update_one({"id": exam_set_id}, {"$set": update_data})

    updated_exam_set = await db.exam_sets.find_one({"id": exam_set_id}, {"_id": 0})
    return updated_exam_set


@api_router.post("/viva-exams/{exam_id}/mark-attendance")
async def mark_attendance(
    exam_id: str,
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can mark attendance"
        )
    
    await db.viva_exams.update_one(
        {"id": exam_id},
        {"$set": {"is_present": True}}
    )
    
    return {"message": "Attendance marked"}


@api_router.get("/viva-exams/my-exams")
async def get_my_exams(current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.STUDENT:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only students can access this"
        )
    
    student_data = await db.students.find_one({"user_id": current_user.id}, {"_id": 0})
    if not student_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student profile not found"
        )  
    
    exams = await db.viva_exams.find({"student_id": student_data["id"]}, {"_id": 0}).to_list(1000)
    
    for exam in exams:
        if exam.get("exam_set_id"):
            exam_set = await db.exam_sets.find_one({"id": exam["exam_set_id"]}, {"_id": 0})
            if exam_set:
                exam["exam_name"] = exam_set.get("name")
                exam["start_time"] = exam_set.get("start_time")
                exam["end_time"] = exam_set.get("end_time")
                exam["duration_minutes"] = exam_set.get("exam_config", {}).get("duration_minutes")
    return exams


@api_router.post("/viva-exams/{exam_id}/start")
async def start_viva_exam(
    exam_id: str,
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.STUDENT:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only students can start exams"
        )
    
    exam_data = await db.viva_exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Exam not found"
        )
    
    exam = VivaExam(**exam_data)
    
    # Attendance check removed in favor of time-based access
    # if not exam.is_present:
    #     raise HTTPException(
    #         status_code=status.HTTP_403_FORBIDDEN,
    #         detail="Attendance not marked. Please contact admin."
    #     )
    
    exam_set_data = await db.exam_sets.find_one({"id": exam.exam_set_id}, {"_id": 0})
    if not exam_set_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Exam set not found"
        )
    
    exam_set = ExamSet(**exam_set_data)
    
    # Handle naive datetimes from DB (local time) vs aware datetimes (UTC)
    if exam_set.start_time.tzinfo is None:
        now = datetime.now()
    else:
        now = datetime.now(timezone.utc)

    if now < exam_set.start_time:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Exam has not started yet"
        )
    
    if now > exam_set.end_time:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The time to start this exam has passed. Please contact your administrator."
        )
    
    # Fetch questions based on subject and topics
    query = {"college_id": exam.college_id}
    if exam_set.selected_subject:
        query["subject"] = exam_set.selected_subject
    if exam_set.selected_topics:
        query["topic"] = {"$in": exam_set.selected_topics}
        
    all_questions = await db.questions.find(query, {"_id": 0}).to_list(2000)
    
    if not all_questions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No questions available for the selected subject/topics"
        )
    
    config = exam_set.exam_config
    selected_questions = []
    
    if config.mode == "topic_wise":
        # Select questions per topic based on rules
        for rule in config.topic_rules:
            topic_qs = [q for q in all_questions if q.get("topic") == rule.topic]
            easy = [q for q in topic_qs if q.get("difficulty_level") == "easy"]
            medium = [q for q in topic_qs if q.get("difficulty_level") == "medium"]
            hard = [q for q in topic_qs if q.get("difficulty_level") == "hard"]
            
            selected_questions.extend(random.sample(easy, max(0, min(rule.easy_count, len(easy)))))
            selected_questions.extend(random.sample(medium, max(0, min(rule.medium_count, len(medium)))))
            selected_questions.extend(random.sample(hard, max(0, min(rule.hard_count, len(hard)))))
    else:
        # Global mode: Select from all filtered questions
        easy = [q for q in all_questions if q.get("difficulty_level") == "easy"]
        medium = [q for q in all_questions if q.get("difficulty_level") == "medium"]
        hard = [q for q in all_questions if q.get("difficulty_level") == "hard"]
        
        selected_questions.extend(random.sample(easy, max(0, min(config.total_easy, len(easy)))))
        selected_questions.extend(random.sample(medium, max(0, min(config.total_medium, len(medium)))))
        selected_questions.extend(random.sample(hard, max(0, min(config.total_hard, len(hard)))))
    
    random.shuffle(selected_questions)
    
    question_ids = [q["id"] for q in selected_questions]
    
    total_questions = len(question_ids)
    max_score = total_questions * 10.0
    
    await db.viva_exams.update_one(
        {"id": exam_id},
        {"$set": {
            "status": ExamStatus.IN_PROGRESS.value,
            "started_at": now.isoformat(),
            "selected_questions": question_ids,
            "max_score": max_score
        }}
    )
    
    return {
        "message": "Exam started",
        "questions": selected_questions,
        "duration_minutes": config.duration_minutes
    }


@api_router.post("/viva-exams/{exam_id}/submit-answer")
async def submit_answer(
    exam_id: str,
    question_id: str = Form(...),
    audio_file: Optional[UploadFile] = File(None),
    transcript: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.STUDENT:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only students can submit answers"
        )
    
    exam_data = await db.viva_exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Exam not found"
        )
    
    question_data = await db.questions.find_one({"id": question_id}, {"_id": 0})
    if not question_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Question not found"
        )
    
    question = Question(**question_data)
    
    student_answer = transcript
    if audio_file and not transcript:
        audio_bytes = await audio_file.read()
        logger.info(f"Received audio file for question {question_id}, size: {len(audio_bytes)} bytes")
        student_answer = await voice_service.transcribe_audio(audio_bytes, "en")
    
    # Ensure it's at least an empty string, not None
    student_answer = student_answer or ""
    
    evaluation = await voice_service.evaluate_answer(
        question=question.question_text,
        student_answer=student_answer,
        answer_key=question.answer_key,
        max_score=10.0
    )
    
    # Check if a response for this question already exists in this exam
    existing_record = await db.qa_records.find_one({
        "viva_exam_id": exam_id,
        "question_id": question_id
    })
    
    # If it's a system action (like repeat or clarify), don't store as an answer
    if evaluation.get("is_system_action", False):
        return {
            "score": 0.0,
            "feedback": evaluation["feedback"],
            "evaluation": evaluation["evaluation"],
            "transcript": student_answer,
            "is_system_action": True,
            "follow_up_hint": ""
        }

    qa_record = QARecord(
        viva_exam_id=exam_id,
        question_id=question_id,
        question_text=question.question_text,
        student_answer=student_answer,
        answer_key=question.answer_key,
        score=evaluation["score"],
        max_score=10.0,
        transcript=student_answer
    )
    
    qa_dict = qa_record.model_dump()
    qa_dict['created_at'] = qa_dict['created_at'].isoformat()
    
    current_total = exam_data.get("total_score", 0.0)
    
    if existing_record:
        # Update existing record
        old_score = existing_record.get("score", 0.0)
        await db.qa_records.update_one(
            {"viva_exam_id": exam_id, "question_id": question_id},
            {"$set": qa_dict}
        )
        new_total = current_total - old_score + evaluation["score"]
    else:
        # Insert new record
        await db.qa_records.insert_one(qa_dict)
        new_total = current_total + evaluation["score"]
    
    await db.viva_exams.update_one(
        {"id": exam_id},
        {"$set": {"total_score": new_total}}
    )
    
    return {
        "score": evaluation["score"],
        "feedback": evaluation["feedback"],
        "evaluation": evaluation["evaluation"],
        "transcript": student_answer,
        "is_system_action": evaluation.get("is_system_action", False),
        "follow_up_hint": evaluation.get("follow_up_hint", "")
    }


@api_router.post("/viva-exams/{exam_id}/complete")
async def complete_exam(
    exam_id: str,
    current_user: User = Depends(get_current_user_dep)
):
    if current_user.role != UserRole.STUDENT:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only students can complete exams"
        )
    
    await db.viva_exams.update_one(
        {"id": exam_id},
        {"$set": {
            "status": ExamStatus.COMPLETED.value,
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Exam completed"}


@api_router.get("/viva-exams/{exam_id}/report")
async def get_exam_report(
    exam_id: str,
    current_user: User = Depends(get_current_user_dep)
):
    # Allow Student (own exam) or College Admin (any exam in their college)
    exam_data = await db.viva_exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam_data:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    if current_user.role == UserRole.STUDENT:
        student = await db.students.find_one({"user_id": current_user.id})
        if not student or student["id"] != exam_data["student_id"]:
            raise HTTPException(status_code=403, detail="Not authorized")
    elif current_user.role == UserRole.COLLEGE_ADMIN:
        if exam_data["college_id"] != current_user.college_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Fetch QA Records
    qa_records = await db.qa_records.find({"viva_exam_id": exam_id}, {"_id": 0}).to_list(1000)
    
    # Fetch Exam Set details for context
    exam_set = await db.exam_sets.find_one({"id": exam_data["exam_set_id"]}, {"_id": 0})
    
    return {
        "exam_details": exam_data,
        "exam_set": exam_set,
        "qa_records": qa_records,
        "total_score": exam_data.get("total_score", 0),
        "max_score": exam_data.get("max_score", 0)
    }


@api_router.post("/voice/synthesize", response_model=Dict[str, str])
async def synthesize_voice(
    interaction: VoiceInteraction,
    current_user: User = Depends(get_current_user_dep)
):
    audio_base64 = await voice_service.synthesize_speech_base64(interaction.text, voice="nova")
    return {"audio": audio_base64, "format": "mp3"}


@api_router.post("/voice/transcribe")
async def transcribe_audio_endpoint(
    audio_file: UploadFile = File(...),
    language: str = Form("en"),
    current_user: User = Depends(get_current_user_dep)
):
    audio_bytes = await audio_file.read()
    transcript = await voice_service.transcribe_audio(audio_bytes, language)
    return {"transcript": transcript}


@api_router.get("/analytics/college")
async def get_college_analytics(current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.COLLEGE_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins can access analytics"
        )
    
    college_data = await db.colleges.find_one({"admin_id": current_user.id}, {"_id": 0})
    if not college_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="College not found"
        )
    
    total_students = await db.students.count_documents({"college_id": college_data["id"]})
    total_questions = await db.questions.count_documents({"college_id": college_data["id"]})
    total_exams = await db.viva_exams.count_documents({"college_id": college_data["id"]})
    completed_exams = await db.viva_exams.count_documents({
        "college_id": college_data["id"],
        "status": ExamStatus.COMPLETED.value
    })
    
    completed_exams_list = await db.viva_exams.find({
        "college_id": college_data["id"],
        "status": ExamStatus.COMPLETED.value
    }, {"_id": 0}).sort("completed_at", -1).to_list(1000)
    
    # Enrich with student names and exam set names
    student_reports = []
    for exam in completed_exams_list:
        student = await db.students.find_one({"id": exam["student_id"]})
        exam_set = await db.exam_sets.find_one({"id": exam["exam_set_id"]})
        max_score_val = exam.get("max_score", 0)
        if not max_score_val:
            num_questions = len(exam.get("selected_questions", []))
            max_score_val = num_questions * 10.0 if num_questions > 0 else 10.0

        student_reports.append({
            "viva_exam_id": exam["id"],
            "exam_set_id": exam.get("exam_set_id"),
            "student_name": student["full_name"] if student else "Unknown",
            "student_id": student["student_id"] if student else "N/A",
            "exam_name": exam_set["name"] if exam_set else "Unknown",
            "subject": exam_set.get("selected_subject", "General") if exam_set else "General",
            "score": exam.get("total_score", 0),
            "max_score": max_score_val,
            "date": exam.get("completed_at")
        })
    
    total_score_sum = sum(r['score'] for r in student_reports)
    total_max_score_sum = sum(r['max_score'] for r in student_reports)

    # Calculate average score normalized to a scale of 10
    avg_score_normalized = (total_score_sum / total_max_score_sum * 10) if total_max_score_sum > 0 else 0
    
    return {
        "total_students": total_students,
        "total_questions": total_questions,
        "total_exams": total_exams,
        "completed_exams": completed_exams,
        "average_score": round(avg_score_normalized, 2),
        "student_reports": student_reports
    }


@api_router.get("/analytics/global")
async def get_global_analytics(current_user: User = Depends(get_current_user_dep)):
    if current_user.role != UserRole.GLOBAL_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only global admin can access global analytics"
        )
    
    total_colleges = await db.colleges.count_documents({})
    active_colleges = await db.colleges.count_documents({"subscription_status": SubscriptionStatus.ACTIVE.value})
    total_students = await db.students.count_documents({})
    total_exams = await db.viva_exams.count_documents({})
    
    return {
        "total_colleges": total_colleges,
        "active_colleges": active_colleges,
        "total_students": total_students,
        "total_exams": total_exams
    }


@api_router.get("/")
async def root():
    return {"message": "College Viva Voice-Agent Platform API"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    # Seed Global Admin if not exists
    existing_admin = await db.users.find_one({"email": ADMIN_EMAIL})
    # Check if admin exists and has the correct role and a full_name
    if not existing_admin or existing_admin.get("role") != "global_admin" or not existing_admin.get("full_name"):
        if existing_admin:
            logger.info(f"Incorrect global admin record found for {ADMIN_EMAIL}. Deleting and recreating.")
            await db.users.delete_one({"email": ADMIN_EMAIL})

        logger.info(f"Creating default global admin: {ADMIN_EMAIL}")
        admin_user = User(
            email=ADMIN_EMAIL,
            password_hash=hash_password(ADMIN_PASSWORD),
            role=UserRole.GLOBAL_ADMIN,
            full_name="System Administrator"
        )
        user_dict = admin_user.model_dump()
        user_dict['created_at'] = user_dict['created_at'].isoformat()
        user_dict['updated_at'] = user_dict['updated_at'].isoformat()
        await db.users.insert_one(user_dict)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
